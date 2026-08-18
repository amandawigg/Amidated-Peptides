"""
Build the interactive MS/MS spectral viewer (Supplemental Data).

Wiggenhorn et al., "Global profiling of peptide amidation to discover
bioactive fragments of the secretome."

Produces a single self-contained HTML file: an offline viewer showing, for every
confirmed peptide, the endogenous MS/MS spectrum mirrored against its 1 uM
synthetic standard, with matched fragment ions annotated.

Inputs
------
1. spectra.json          extracted MS/MS scans (see schema below)
2. Supplemental_Table_2  peptide names, fingerprint ions, scores, detection calls
3. viewer_template.html  the viewer interface, with a __SPECTRA_DATA__ placeholder

Fragment ion m/z values and the set of ions to annotate are computed here.

spectra.json schema
-------------------
    {
      "<PEP name>": {
        "standard": {"mz": [...], "intensity": [...],
                     "n_scans": int, "mean_rt": float,
                     "precursor_mz": float, "charge": int},
        "tissues": {
          "<Tissue>": {
            "<replicate number>": {"mz": [...], "intensity": [...],
                                   "n_scans": int, "mean_rt": float,
                                   "precursor_mz": float, "charge": int}
          }
        }
      }
    }

Intensities are normalised to the base peak of each spectrum. Producing this
file requires the raw acquisitions, deposited on Mendeley Data
(DOI: 10.17632/9wm5x5ncfb.1).

Usage
-----
    python build_viewer.py \\
        --spectra  data/spectra.json \\
        --table    data/Supplemental_Table_2.xlsx \\
        --template viewer_template.html \\
        --output   viewer_MSMS.html
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

# ── Monoisotopic residue masses ──────────────────────────────────────────────
RESIDUE = {
    "G":  57.02146, "A":  71.03711, "S":  87.03203, "P":  97.05276,
    "V":  99.06841, "T": 101.04768, "C": 103.00919, "L": 113.08406,
    "I": 113.08406, "N": 114.04293, "D": 115.02694, "Q": 128.05858,
    "K": 128.09496, "E": 129.04259, "M": 131.04049, "H": 137.05891,
    "F": 147.06841, "R": 156.10111, "Y": 163.06333, "W": 186.07931,
}
H2O, PROTON, NH3, CO = 18.0105646, 1.00727646, 17.0265491, 27.9949146
CARBAMIDOMETHYL = 57.0214640          # iodoacetamide adduct on cysteine
AMIDATION       = -0.98402            # C-terminal -OH replaced by -NH2
PYROGLUTAMATE   = -NH3                # N-terminal Gln cyclisation

MAX_CHARGE       = 3
N_ANNOTATED_IONS = 10   # most intense ions in the standard that are labelled,
                        # in addition to the fingerprint ions


def parse_sequence(seq: str) -> list[float]:
    """Residue masses for a modified sequence, e.g. 'pGlu-C*RPSK-NH2'."""
    seq = seq.replace("₂", "2").strip()
    masses: list[float] = []

    if seq.startswith("pGlu-"):
        masses.append(RESIDUE["Q"] + PYROGLUTAMATE)
        seq = seq[5:]
    seq = seq.replace("-NH2", "")

    i = 0
    while i < len(seq):
        aa = seq[i]
        i += 1
        m = RESIDUE[aa]
        if i < len(seq) and seq[i] == "*":      # carbamidomethyl-Cys
            m += CARBAMIDOMETHYL
            i += 1
        masses.append(m)
    return masses


def fragment_ions(seq: str) -> dict[str, float]:
    """
Theoretical m/z for all a, b, y and y-NH3 ions up to MAX_CHARGE."""
    res = parse_sequence(seq)
    n = len(res)
    neutral: dict[str, float] = {}

    running = 0.0                                   # N-terminal series
    for i in range(1, n):
        running += res[i - 1]
        neutral[f"b{i}"] = running
        neutral[f"a{i}"] = running - CO

    c_term = H2O + AMIDATION                        # C-terminal series
    running = 0.0
    for i in range(1, n):
        running += res[n - i]
        y = running + c_term
        neutral[f"y{i}"] = y
        neutral[f"y{i}-NH3"] = y - NH3

    out: dict[str, float] = {}
    for name, mass in neutral.items():
        for z in range(1, MAX_CHARGE + 1):
            label = name if z == 1 else f"{name}+{z}"
            out[label] = (mass + z * PROTON) / z
    return out


# ── Supplemental Table 2 ─────────────────────────────────────────────────────
def read_table(path: Path) -> dict:
    """Peptide metadata, fingerprint ions and per-replicate scores."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    rows = [list(r) for r in wb["Detected Peptides"].iter_rows(values_only=True)]
    header = [str(c) if c else "" for c in rows[1]]
    col = {name: i for i, name in enumerate(header)}
    peptides = {
        r[col["PEP Name"]]: {
            "sequence":    r[col["Peptide Sequence"]],
            "fingerprint": [i.strip() for i in str(r[col["Ions for Dot Product"]]).split(",")],
        }
        for r in rows[2:] if r and r[col["PEP Name"]]
    }

    rows = [list(r) for r in wb["spectral valid"].iter_rows(values_only=True)]
    header = [str(c) if c else "" for c in rows[1]]
    col = {name: i for i, name in enumerate(header)}
    scores: dict[tuple[str, str], dict] = {}
    for r in rows[2:]:
        if not r or not r[col["PEP Name"]]:
            continue
        scores[(r[col["PEP Name"]], str(r[col["Replicate"]]))] = {
            "dot_product":       r[col["Dot Product"]],
            "ppm_error":         r[col["Precursor Accuracy (ppm)"]],
            "ion_count":         r[col["Ion Count"]],
            "is_real_detection": str(r[col["Is Detected"]]).strip() == "Yes",
        }

    wb.close()
    return {"peptides": peptides, "scores": scores}


# ── Assembly ─────────────────────────────────────────────────────────────────
def annotated_ions(spectrum: dict, theoretical: dict[str, float],
                   fingerprint: list[str], tol: float = 0.8) -> dict[str, float]:
    """
The ions labelled in the plots: the N_ANNOTATED_IONS most intense in the
    synthetic standard, plus every fingerprint ion. Unmatched peaks render as
    background.
    """
    def intensity(mz: float) -> float:
        return max((i for m, i in zip(spectrum["mz"], spectrum["intensity"])
                    if abs(m - mz) <= tol), default=0.0)

    ranked = sorted(theoretical, key=lambda ion: -intensity(theoretical[ion]))
    keep = set(ranked[:N_ANNOTATED_IONS]) | set(fingerprint)
    return {ion: theoretical[ion] for ion in keep if ion in theoretical}


def build(spectra: dict, table: dict) -> dict:
    data: dict[str, dict] = {}

    for name, meta in table["peptides"].items():
        if name not in spectra:
            print(f"  no spectra for {name}, skipped")
            continue

        entry = spectra[name]
        theoretical = fragment_ions(meta["sequence"])
        frag_map = annotated_ions(entry["standard"], theoretical, meta["fingerprint"])

        tissues: dict[str, dict] = {}
        for tissue, reps in entry["tissues"].items():
            tissues[tissue] = {}
            for rep, spec in reps.items():
                score = table["scores"].get((name, f"{tissue}_{rep}"), {})
                tissues[tissue][rep] = {**spec, **score}

        data[name] = {
            "pep_name":     name,
            "sequence":     meta["sequence"],
            "top4_ions":    meta["fingerprint"],
            "frag_mz_map":  frag_map,
            "standard":     entry["standard"],
            "tissues":      tissues,
        }

    return data


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spectra",  type=Path, default=Path("../data/spectra.json"))
    p.add_argument("--table",    type=Path, default=Path("../data/Supplemental_Table_2.xlsx"))
    p.add_argument("--template", type=Path, default=Path("viewer_template.html"))
    p.add_argument("--output",   type=Path, default=Path("viewer_MSMS.html"))
    args = p.parse_args()

    for path in (args.spectra, args.table, args.template):
        if not path.exists():
            p.error(f"missing input: {path}")

    spectra = json.loads(args.spectra.read_text())
    table   = read_table(args.table)
    print(f"peptides in table: {len(table['peptides'])}"
          f" | with spectra: {len(spectra)}")

    data = build(spectra, table)

    template = args.template.read_text(encoding="utf8")
    if "__SPECTRA_DATA__" not in template:
        p.error("template has no __SPECTRA_DATA__ placeholder")
    html = template.replace("__SPECTRA_DATA__",
                            json.dumps(data, separators=(",", ":")))
    args.output.write_text(html, encoding="utf8")

    detections = sum(r.get("is_real_detection", False)
                     for d in data.values()
                     for reps in d["tissues"].values()
                     for r in reps.values())
    print(f"peptides written : {len(data)}")
    print(f"detections shown : {detections}")
    print(f"output           : {args.output} "
          f"({args.output.stat().st_size / 1e6:.1f} MB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
